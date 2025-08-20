import pandas as pd
from pandas import DataFrame, Series
from pandas._libs.tslibs.timestamps import Timestamp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from typing import Optional, List
import matplotlib.pyplot as plt



def get_avg_and_median_if_condition(
    df: DataFrame,
    numeric_column: str,
    condition_column: str,
    condition_value: str = "PRESERVED",
    inverted: Optional[bool] = False
) -> dict:
    """
    Returns the average and median of a numeric column, filtered by a condition in another column.

    Args:
        df (DataFrame): Input DataFrame.
        numeric_column (str): Name of the numeric column to analyze.
        condition_column (str): Column used for filtering (e.g. 'Status').
        condition_value (str): Value to filter on in the condition column (default is 'PRESERVED').

    Returns:
        dict: A dictionary with average and median of the filtered values.
    """
    # Clean column names
    condition_series = df[condition_column].astype(str).str.strip().str.upper()
    target_value = condition_value.strip().upper()

    if inverted:
        filtered = df[condition_series != target_value]
    else:
        filtered = df[condition_series == target_value]
    # Convert the target column to numeric (in case it's stored as string)
    values = pd.to_numeric(filtered[numeric_column], errors='coerce').dropna()

    if values.empty:
        return {'average': None, 'median': None}

    return {
        'average': values.mean(),
        'median': values.median()
    }

def sum_column_if_condition(
    df: DataFrame,
    numeric_column: str,
    condition_column: str,
    condition_value: str = "PRESERVED",
    invert: bool = False
) -> float:
    """
    Returns the sum of a numeric column, filtered by a condition in another column.

    Args:
        df (DataFrame): Input DataFrame.
        numeric_column (str): Name of the numeric column to sum.
        condition_column (str): Column used for filtering (e.g. 'Status').
        condition_value (str): Value to filter on in the condition column (default is 'PRESERVED').
        invert (bool): If True, sums rows NOT matching the condition.

    Returns:
        float: The total sum of the numeric column after filtering.
    """
    df.columns = df.columns.str.strip()

    # Clean and normalize condition column
    condition_series = df[condition_column].astype(str).str.strip().str.upper()
    target_value = condition_value.strip().upper()

    # Apply filter
    if invert:
        filtered = df[condition_series != target_value]
    else:
        filtered = df[condition_series == target_value]

    # Convert and sum
    values = pd.to_numeric(filtered[numeric_column], errors='coerce').dropna()
    return values.sum()




def highlight_non_blank_cells(
    excel_path: str,
    sheet_name: str,
    column_letter: str,
    highlight_color: str = "FFFF00"  # Yellow
) -> None:
    """
    Highlights non-blank cells in a specified column of an Excel sheet.

    Args:
        excel_path (str): Path to the Excel workbook.
        sheet_name (str): Name of the worksheet to modify.
        column_letter (str): Excel-style column letter (e.g., 'E').
        highlight_color (str): Hex color code for highlighting. Default is yellow.
    """
    totalHighlighted = 0
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    
    ws = wb[sheet_name]
    fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    for row in range(2, ws.max_row + 1):  # Skipping header row
        cell = ws[f"{column_letter}{row}"]
        if cell.value not in (None, "", " ", "NA", "N/A"):
            totalHighlighted += 1
            cell.fill = fill

    wb.save(excel_path)
    print(f"Highlighted non-blank cells in column {column_letter} on sheet '{sheet_name}'.")
    print(f"Total highlighted cells: {totalHighlighted}")



def mark_preserved_status_by_section8(
    excel_path: str,
    sheet_name: str,
    forecast_date: str,
    status_column_name: str = "Status"
) -> None:
    """
    Updates the Status column to 'PRESERVED' for rows where:
    - S8_ExpDate > forecast_date
    - S8_PBAUnits == TotalUnits
    Only updates values in Excel — keeps all formatting intact.

    Args:
        excel_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to modify.
        forecast_date (str): Forecast cutoff date in MM/DD/YYYY format.
        status_column_name (str): The column to update (default is 'Status').
    """
    # Load data with pandas (only for logic)
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.columns = df.columns.str.strip()

    # Convert date and numeric types
    df['S8_ExpDate'] = pd.to_datetime(df['S8_ExpDate'], errors='coerce')
    df['TotalUnits'] = pd.to_numeric(df['TotalUnits'], errors='coerce')
    df['S8_PBAUnits'] = pd.to_numeric(df['S8_PBAUnits'], errors='coerce')

    # Determine rows that meet condition
    cutoff = pd.to_datetime(forecast_date)
    condition = (df['S8_ExpDate'] > cutoff) & (df['S8_PBAUnits'] == df['TotalUnits'])

    # Get index of rows to update (add 2 because openpyxl is 1-based and assumes header in row 1)
    rows_to_update = df[condition].index + 2

    # Load workbook and sheet with openpyxl
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Find the Excel column index for the "Status" column
    header = [cell.value for cell in ws[1]]
    try:
        col_idx = header.index(status_column_name) + 1
    except ValueError:
        raise ValueError(f"Column '{status_column_name}' not found in Excel sheet.")

    # Update only the needed cells
    for row in rows_to_update:
        ws.cell(row=row, column=col_idx).value = "PRESERVED"

    # Save
    wb.save(excel_path)
    print(f"Updated 'Status' to 'PRESERVED' for {len(rows_to_update)} properties.")





if __name__ == "__main__":
    
    
    print("Script Started")
    df: DataFrame = pd.read_excel("2030 Forecast E-U-R Draft -Local.xlsx", sheet_name= "All 202 Properties", na_values=["", " ", "  ", "NA", "N/A", "null"])
    df.columns = df.columns.str.strip()


    
    
    

