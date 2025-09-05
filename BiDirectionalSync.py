import pandas as pd
from pandas import DataFrame, Series
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

import re
from typing import Optional, List
import matplotlib.pyplot as plt


def load_sheet_as_df(excel_path: str, sheet_name: str, id_column: str) -> Optional[DataFrame]:
    """Load and clean an Excel sheet into a DataFrame indexed by the ID column.
    Returns None if ID column is missing.
    """
    df = pd.read_excel(excel_path, sheet_name=sheet_name).fillna("")
    df.columns = df.columns.str.strip()


    if id_column not in df.columns:
        print(f"⚠️ Skipping '{sheet_name}': Missing required ID column '{id_column}'")
        return None

    df[id_column] = df[id_column].astype(str).str.strip()
    df.set_index(id_column, inplace=True)
    return df


def append_unique_values(existing: str, new: str, separator: str = " | ") -> str:
    """Append new to existing if it's not already included."""
    if not existing:
        return new
    existing_parts = [part.strip() for part in existing.split(separator)]
    if new.strip() in existing_parts or not new.strip():
        return existing
    return separator.join(existing_parts + [new.strip()])

def update_main_from_subsheet(df_main: DataFrame, df_sub: Optional[DataFrame], sync_columns: List[str], sheet_name: str) -> None:
    """Update the main DataFrame by appending unique values from a subsheet."""
    if not all(col in df_sub.columns for col in sync_columns):
        print(f"⚠️ Skipping '{sheet_name}': Missing one or more required columns: {sync_columns}")
        return

    for prop_id in df_sub.index:
        if prop_id not in df_main.index:
            continue
        for col in sync_columns:
            sub_val = df_sub.at[prop_id, col].strip()
            main_val = df_main.at[prop_id, col].strip()
            updated_val = append_unique_values(main_val, sub_val)
            df_main.at[prop_id, col] = updated_val


def update_subsheet_from_main(df_sub: DataFrame | None, df_main: DataFrame, sync_columns: List[str], sheet_name: str) -> None:
    """Overwrite subsheet values with values from the main sheet."""
    if not all(col in df_sub.columns for col in sync_columns):
        print(f"⚠️ Skipping '{sheet_name}': Missing one or more required columns: {sync_columns}")
        return

    for prop_id in df_sub.index:
        if prop_id not in df_main.index:
            continue
        for col in sync_columns:
            df_sub.at[prop_id, col] = df_main.at[prop_id, col]


def write_df_to_worksheet(ws: Worksheet, df: DataFrame) -> None:
    """Write a DataFrame to a worksheet, including headers."""
    df_reset = df.reset_index()
    for col_idx, header in enumerate(df_reset.columns, start=1):
        ws.cell(row=1, column=col_idx).value = header

    for row_idx, (_, row_data) in enumerate(df_reset.iterrows(), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

def sync_columns_bidirectional(
    excel_path: str,
    id_column: str = "PropertyID",
    sync_columns: List[str] = ["Comments", "Status"],
    main_sheet: str = "Main forecast"
) -> None:
    """Orchestrates syncing columns between Main sheet and subsheets."""
    wb = load_workbook(excel_path)
    all_sheets = wb.sheetnames

    if main_sheet not in all_sheets:
        raise ValueError(f"Main sheet '{main_sheet}' not found in workbook.")

    # Load Main forecast
    df_main = load_sheet_as_df(excel_path, main_sheet, id_column)

    # 1. Update Main from subsheets
    for sheet in all_sheets:
        if sheet == main_sheet:
            continue
        print(f"🔄 Subsheet → Main: {sheet}")
        df_sub = load_sheet_as_df(excel_path, sheet, id_column)
        if df_sub is None:
            continue
        update_main_from_subsheet(df_main, df_sub, sync_columns, sheet)

    # 2. Write updated Main forecast
    print("💾 Writing updated Main forecast...")
    write_df_to_worksheet(wb[main_sheet], df_main)

    # 3. Push updates from Main → SubSheets
    for sheet in all_sheets:
        if sheet == main_sheet:
            continue
        print(f"⬇️ Main → Subsheet: {sheet}")
        df_sub = load_sheet_as_df(excel_path, sheet, id_column)
        if df_sub is None:
            continue
        update_subsheet_from_main(df_sub, df_main, sync_columns, sheet)
        write_df_to_worksheet(wb[sheet], df_sub)

    # Save all updates
    wb.save(excel_path)
    print("\n✅ Bidirectional sync complete.")

if __name__ == "__main__":
        
    print("Script Started")
    #f: DataFrame = pd.read_excel("2030 Forecast E-U-R Draft -Local.xlsx", sheet_name= "All 202 Properties", na_values=["", " ", "  ", "NA", "N/A", "null"])
    #df.columns = df.columns.str.strip()

    print("Dataframe Loaded.")
    
    update_subsheet_from_main = 
    sync_columns_bidirectional(
        excel_path="2030 Forecast E-U-R Draft -Master.xlsx",
        id_column="PropertyID",
        sync_columns=["1st Cut Status", "Jameer notes", "Roberta notes"],
        main_sheet = "Main Forecast"
        )
    