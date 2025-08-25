import pandas as pd
from pandas import DataFrame, Series
from pandas._libs.tslibs.timestamps import Timestamp
from openpyxl import load_workbook
import re
from typing import Optional



# This file contains code for running various checks for the CEDAC At-Risk Forecast report. It was built for the 2030 report, but contains functions that can be used for other years.
# Written by Jameer Gomez-Santos, jgomez-santos@cedac.org, 2025


# Change this to the desired forecast date in MM/DD/YYYY format
forecast_date: str = "12/31/2030"
print("Script Started")
df: DataFrame = pd.read_excel("advance Preserve Through date to 2030 v3 draft.xlsx", na_values=["", " ", "  ", "NA", "N/A", "null"])
df.columns = df.columns.str.strip()

print("Dataframe Loaded.")



def check_section_8_expiry(df: DataFrame, forecast_date: str) -> DataFrame:
    '''Returns a Dataframe of all properties from the inputed dataframe that consist of 100% S8 units and have an S8 expiry date after the forcast date.

    Args:
        df (DataFrame): Dataframe all the forecast data. Must have the columns 'S8_ExpDate', 'S8_PBAUnits', and 'TotalUnits'.
        forecast_date (str): The cutoff date for the forecast in MM/DD/YYYY format.

    Returns:
        DataFrame: Subset of input Dataframe containing properties with 100% Section 8 units and an S8 expiry date after the forecast date.
    '''
    # Convert the forecast date to a datetime object
    cutOff_DT:Timestamp = pd.to_datetime(forecast_date)
    # Ensure the S8_ExpDate column is in datetime format
    df['S8_ExpDate'] = pd.to_datetime(df['S8_ExpDate'], errors='coerce')
    df['TotalUnits'] = pd.to_numeric(df['TotalUnits'], errors='coerce').fillna(0).astype(int)
    df['S8_PBAUnits'] = pd.to_numeric(df['S8_PBAUnits'], errors='coerce')
    # Filter for properties that are 100% Section 8 units and have an expiry date after the forecast date
    section_8_expiry:DataFrame = df[(df['S8_ExpDate'] > cutOff_DT) & (df['S8_PBAUnits'] == df['TotalUnits'])]
    return section_8_expiry


def get_all_s8_properties(df: DataFrame) -> DataFrame:
    """
    Returns a DataFrame of all properties that have at least one S8_PBAUnit.

    Args:
        df (DataFrame): The input DataFrame containing property data, including the 'S8_PBAUnits' column.

    Returns:
        DataFrame: Subset of the input DataFrame where S8_PBAUnits is not null and greater than 0.
    """
    df['S8_PBAUnits'] = pd.to_numeric(df['S8_PBAUnits'], errors='coerce')
    s8_properties = df[df['S8_PBAUnits'].notna() & (df['S8_PBAUnits'] > 0)]
    return s8_properties

def find_all_overrides(df: DataFrame, forecast_date: Optional[str] = None) -> DataFrame:
    """
    Returns a DataFrame of all properties with an override date.
    If a forecast_date is provided, only properties with override dates after that date are included.

    Args:
        df (DataFrame): Input DataFrame containing property data.
        forecast_date (Optional[str]): Optional cutoff date (MM/DD/YYYY) for filtering override dates.

    Returns:
        DataFrame: Subset of DataFrame with 'PropertyID' and 'OverrideDate'.
    """
    pattern = r"At Risk Date Override:\s*(\d{1,2}/\d{1,2}/\d{4})"
    df['OverrideDate'] = df['ForecastPreservedByProgram'].astype(str).str.extract(pattern)
    df['OverrideDate'] = pd.to_datetime(df['OverrideDate'], errors='coerce')

    if forecast_date:
        cutoff_dt = pd.to_datetime(forecast_date)
        overrides = df[df['OverrideDate'] > cutoff_dt][['PropertyID', 'OverrideDate']]
    else:
        overrides = df[df['OverrideDate'].notna()][['PropertyID', 'OverrideDate']]
    
    return overrides

def extract_propertyIDs(df: DataFrame) -> Series:
    """
    Extracts unique PropertyIDs from the DataFrame.

    Args:
        df (DataFrame): The input DataFrame containing property data.

    Returns:
        Series: A Series of unique PropertyIDs. A series is an individual column of a Dataframe.
    """
    return df['PropertyID']



def get_larger_date(date1: str, date2: str) -> Optional[str]:
    """
    Returns the later of two dates in MM/DD/YYYY format.
    If both dates are invalid or missing, returns None.
    """
    dateTime1 = pd.to_datetime(date1, errors='coerce')
    dateTime2 = pd.to_datetime(date2, errors='coerce')

    if pd.isna(dateTime1) and pd.isna(dateTime2):
        return None
    if pd.isna(dateTime1):
        return dateTime2.strftime('%m/%d/%Y')  # type: ignore
    if pd.isna(dateTime2):
        return dateTime1.strftime('%m/%d/%Y')

    return max(dateTime1, dateTime2).strftime('%m/%d/%Y')

def S8_expDate_vs_overrideDate(S8Propertys: DataFrame, overridenProperties: DataFrame) -> DataFrame:
    """
    Returns a DataFrame of properties where the Section 8 expiration date is later than the override date.

    Args:
        S8Propertys (DataFrame): DataFrame with at least 'PropertyID' and 'S8_ExpDate' columns.
        overridenProperties (DataFrame): DataFrame with 'PropertyID' and 'OverrideDate'.

    Returns:
        DataFrame: Subset of merged DataFrame where S8_ExpDate > OverrideDate.
    """
    # Merge on PropertyID
    merged: DataFrame = pd.merge(S8Propertys, overridenProperties, on="PropertyID", how="inner")

    # Ensure both date columns are in datetime format
    merged['S8_ExpDate'] = pd.to_datetime(merged['S8_ExpDate'], errors='coerce')
    merged['OverrideDate'] = pd.to_datetime(merged['OverrideDate'], errors='coerce')

    # Filter where Section 8 Expiration Date is later than the override date
    result: DataFrame = merged[merged['S8_ExpDate'] > merged['OverrideDate']]

    return result

def extract_Column_With_IDs(df: DataFrame, columnName: str) -> DataFrame:
    '''Extracts single given column along with another column containing its 'PropertyID' from a DataFrame object. Assumes existance of "PropertyID" column.

    Args:
        df (DataFrame): DataFrame to extract a coulmn from.
        columnName (str): Exact name of the coulumn to extract.

    Raises:
        ValueError: Returns a ValueError if the coulumn is not found in the DataFrane.

    Returns:
        Series: A singlar column of the dataframe.
    '''
    if columnName not in df.columns:
        raise ValueError(f"Column '{columnName}' not found in DataFrame.")
    else:
         return df[['PropertyID', columnName]].copy()

def get_Expired_S8(df: DataFrame, forecast_date: str) -> DataFrame:
    """
    Returns full rows of all properties where the S8_ExpDate is earlier than the forecast date (i.e., already expired).

    Args:
        df (DataFrame): The input DataFrame containing property data.
        forecast_date (str): The forecast cutoff date in MM/DD/YYYY format.
 
    Returns:
        DataFrame: Subset of original DataFrame with all columns, for expired S8 properties.
    """
    # Convert columns to proper types
    df['S8_ExpDate'] = pd.to_datetime(df['S8_ExpDate'], errors='coerce')
    
    # Convert forecast_date to datetime
    forecast_dt = pd.to_datetime(forecast_date, errors='coerce')
    
    # Filter full DataFrame
    expired_df = df[df['S8_ExpDate'] < forecast_dt]
    
    return expired_df


expiredS8 = get_Expired_S8(df, forecast_date)
expiredS8.to_excel("Expired_S8_Properties.xlsx", index=False)
'''
matching_ids: Series = extract_propertyIDs(check_section_8_expiry(df, forecast_date))
# Get override properties with their override dates
override_df: DataFrame = find_all_overrides(df)
print(f"📝 Overrides detected for {len(override_df)} properties.")  # <-- This line prints the count


# Update notes and status for overrides
for _, row in override_df.iterrows():
    pid = row['PropertyID']
    override_date = row['OverrideDate']
    
    # Build override note
    override_note = f"Has an override of {override_date}"
    
    # Append to existing note if present
    existing_note = df.loc[df['PropertyID'] == pid, 'Jameer notes'].astype(str).values[0]
    if pd.notna(existing_note) and existing_note.strip() != 'nan':
        new_note = f"{existing_note}; {override_note}"
    else:
        new_note = override_note
    
    df.loc[df['PropertyID'] == pid, 'Jameer notes'] = new_note
    df.loc[df['PropertyID'] == pid, 'Status after 1st cut (done, small, <>)'] = 'done'

note = '100% of units are S.8, and S8 expires after forecast date'
status = 'done'
df.loc[df['PropertyID'].isin(matching_ids), 'Jameer notes'] = note
df.loc[df['PropertyID'].isin(matching_ids), 'Status after 1st cut (done, small, <>)'] = status


# Load the workbook and active sheet with openpyxl (preserves formatting)
workbook = load_workbook("advance Preserve Through date to 2030 v3 draft.xlsx")
sheet = workbook.active

# Get header row to find column indexes
header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
prop_id_idx = header.index("PropertyID")
note_idx = header.index("Jameer notes")
status_idx = header.index("Status after 1st cut (done, small, <>)")

matching_counter = 0
# Iterate through rows and update notes/status for matching IDs
for row in sheet.iter_rows(min_row=2):  # assuming headers are in row 1
    property_id = row[prop_id_idx].value
    if property_id in matching_ids.values:
        matching_counter += 1
        row[note_idx].value = note
        row[status_idx].value = status
    
# Save to a new file (original formatting preserved)
workbook.save("S8_Forecast_Check_Formatted.xlsx")
print(f"\n✅ {matching_counter} properties updated in the Excel file.")
print(f"📊 Total properties with non-null S8_PBAUnits: {find_all_s8_properties(df)}")

'''