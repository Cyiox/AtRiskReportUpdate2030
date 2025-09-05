import pandas as pd
import openpyxl
from copy import copy  # Add this import at the top




def standardize_property_id_column(df: pd.DataFrame) -> pd.DataFrame:
    if "ExpUsePropertyID" in df.columns and "PropertyID" not in df.columns:
        df = df.rename(columns={"ExpUsePropertyID": "PropertyID"})
    return df


def read_All_Sources() -> dict:
    key_map:dict = {
     "DHCD": "ExpUseSources/DHCD_archive.xlsx",
     "LIHTC": "ExpUseSources/LIHTCPUB_archive.xlsx",
     "HUDContract": "ExpUseSources/MF_Assistance_and_Sec8_Contracts_archive.xlsx",
     "HUDProperty": "ExpUseSources/MF_Properties_with_Assistance_and_Sec8_Contracts_archive.xlsx",
     "Mortgage": "ExpUseSources/mtg_a_archive.xlsx",
     "Properties": "ExpUseSources/PropertiesTable.xlsx",
     "MainForecast": "2030 Forecast E-U-R Draft -Local.xlsx"

}

    dataframes:dict = {}
    for key, filename in key_map.items():
     try:
        df = pd.read_excel(filename)
        df = standardize_property_id_column(df)  # Standardize column
        dataframes[key] = df
        print(f"{key} Has been loaded")
     except Exception as e:
            print(f"Error loading {filename}: {e}")
    return dataframes

def keep_latest_import(df: pd.DataFrame, key: str) -> pd.DataFrame:
    if "importID" not in df.columns and "ImportID" not in df.columns:
        raise ValueError(f"DataFrame '{key}' is missing an 'importID' or 'ImportID' column")

    # Normalize the column name if it's capitalized differently
    if "ImportID" in df.columns and "importID" not in df.columns:
        df = df.rename(columns={"ImportID": "importID"})

    if "PropertyID" not in df.columns:
        raise ValueError(f"DataFrame '{key}' is missing a 'PropertyID' column")

    idx = df.groupby("PropertyID")["importID"].idxmax()
    return df.loc[idx].reset_index(drop=True)

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
import pandas as pd


def apply_formatting_from_old_workbook(old_path, new_path):
    from openpyxl import load_workbook
    
    old_wb = load_workbook(old_path)
    new_wb = load_workbook(new_path)
    
    for sheet_name in old_wb.sheetnames:
        if sheet_name not in new_wb.sheetnames:
            continue
        old_ws = old_wb[sheet_name]
        new_ws = new_wb[sheet_name]
        
        # Copy freeze panes
        new_ws.freeze_panes = old_ws.freeze_panes
        
        # Copy autofilter
        if old_ws.auto_filter.ref:
            new_ws.auto_filter.ref = old_ws.auto_filter.ref
        
        # Copy styles within used range
        max_row = old_ws.max_row
        max_col = old_ws.max_column
        
        for row in old_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.col_idx)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
    
    new_wb.save(new_path)
    print("Styles Changed")

if __name__ == "__main__":
    

    apply_formatting_from_old_workbook("2030 Forecast E-U-R Draft -Local.xlsx", "MergedForecast.xlsx")

    '''
    dataframes = read_All_Sources()
    print("Data read succesfully")
    for key in ["DHCD", "HUDContract","HUDProperty"]:
        df = dataframes[key]
        dataframes[key] = keep_latest_import(df,key)
    print("Latest imports grabbed")


    selected_cols = {
    "Properties": ["PropertyID", "ProjectID_lending","Comment: Preservation Status, Type or other new Affordability Te","Mortgage status/ TC and other restrictions","Project/AKA"],
    "HUDProperty": ["PropertyID", "owner_email_text"],
    "HUDContract": ["PropertyID","tracs_effective_date"],
    "DHCD": ["PropertyID", "CBH Awarded", "HIF Awarded","FCF DDS Awarded","FCF DMH Awarded","Date CBH Loan Closed","Date HIF Loan Closed","Date FCF DDS Loan Closed","Date FCF DMH Loan Closed"]
    }

    base_df = None

    for filename, cols in selected_cols.items():
        df = dataframes[filename][cols]
    
        # Rename non-ID columns to include source tag
        renamed_cols = {col: f"{filename.split('_')[0]}_{col}" for col in cols if col != "PropertyID"}
        df = df.rename(columns=renamed_cols)

        if base_df is None:
         base_df = df
        else:
         base_df = base_df.merge(df, on="PropertyID", how="left")

    # Load forecast workbook (all sheets)
    forecast_sheets = pd.read_excel("2030 Forecast E-U-R Draft -Local.xlsx", sheet_name=None)

    # Ensure PropertyID is string in base_df
    base_df["PropertyID"] = base_df["PropertyID"].astype(str)

    # Dictionary to hold updated sheets
    updated_sheets = {}

    # Process each sheet
    for sheet_name, df in forecast_sheets.items():
        # Skip Forecast notes or any sheet without PropertyID
        if sheet_name == "Forecast notes":
            print(f"Skipping sheet '{sheet_name}' – notes only.")
            updated_sheets[sheet_name] = df
            continue

        # Standardize PropertyID column
        if "ExpUsePropertyID" in df.columns and "PropertyID" not in df.columns:
            df = df.rename(columns={"ExpUsePropertyID": "PropertyID"})

        if "PropertyID" not in df.columns:
            print(f"Skipping sheet '{sheet_name}' – no PropertyID column found.")
            updated_sheets[sheet_name] = df
            continue

        # Ensure same type for merge
        df["PropertyID"] = df["PropertyID"].astype(str)

        # Merge base_df into this sheet
        merged = df.merge(base_df, on="PropertyID", how="left")
        updated_sheets[sheet_name] = merged
        print(f"Merged base data into '{sheet_name}': {merged.shape[0]} rows")

    # Write all updated sheets to a new Excel file
    with pd.ExcelWriter("MergedForecast.xlsx", engine="openpyxl") as writer:
        for name, sheet_df in updated_sheets.items():
            sheet_df.to_excel(writer, sheet_name=name, index=False)

    print("✅ All forecast sheets successfully merged and saved to 'MergedForecast.xlsx'")
    '''
    